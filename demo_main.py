#!/usr/bin/python
# -*- coding: utf-8 -*-

########################################################################################################################
#
# Copyright (c) 2023 YangZhengli All Rights Reserved.
# File              :   DataQuery.py
# Description       :   DataQuery
# Date              :   2023-03-18
# Owner             :   zhengli.yang@boschhuayu-steering.com
# History           :   2023-03-18:First created
#
########################################################################################################################

from logging_maker import logger

print("*" * 80)
print("SysEng 2.0 Data Query Tool")
print("Version: demo V1.1")
print("Author: zhengli.yang@boschhuayu-steering.com")
print("*" * 80 + "\n")

logger.info("SysEng 2.0 Data Query Tool")
logger.info("Version: demo V1.1")
logger.info("开始运行")


# from get_file_path import *
import xlrd
from openpyxl.styles import PatternFill
from compare_text import simp_str
from find_value import variable_name_list,dict_mech_parameter,dict_a2l_parameter,parameter_list,fun_find_para
from find_text import format_paragraph_find
import openpyxl
import os
import time



def check_title(first_row):
    check_title_result = False
    data_module_pos_col_result = False
    wert_pos_col_result = False
    tuning_pra_pos_col_result = False
    data_module_pos_col = 0
    wert_pos_col = 0
    tuning_pra_pos_col = 0
    for i in range(len(first_row)):
        # print(f"first_row[i] = {first_row[i]}")
        if not data_module_pos_col_result:
            if "Data Module" in first_row[i]:
                # print(i)
                data_module_pos_col = i
                data_module_pos_col_result = True
                print(f"Data Module在第{data_module_pos_col + 1}列")
                logger.info(f"Data Module在第{data_module_pos_col + 1}列")
            else:
                data_module_pos_col_result = False
        if not wert_pos_col_result:
            if "Wert(e)" in first_row[i]:
                # print(i)
                wert_pos_col = i
                wert_pos_col_result = True
                print(f"Wert(e)在第{wert_pos_col + 1}列")
                logger.info(f"Wert(e)在第{wert_pos_col + 1}列")
            else:
                wert_pos_col_result = False
        if not tuning_pra_pos_col_result:
            if "Tuning Parameter" in first_row[i]:
                # print(i)
                tuning_pra_pos_col = i
                tuning_pra_pos_col_result = True
                print(f"Tuning Parameter在第{tuning_pra_pos_col + 1}列")
                logger.info(f"Tuning Parameter在第{tuning_pra_pos_col + 1}列")
            else:
                tuning_pra_pos_col_result = False
    if data_module_pos_col_result and wert_pos_col_result and tuning_pra_pos_col_result:
        check_title_result = True
    else:
        check_title_result = False
        print("表格不完整")
        logger.critical("表格不完整")
    # print(f"check_title_result = {check_title_result})
    return check_title_result, data_module_pos_col, wert_pos_col, tuning_pra_pos_col


def update_wert(file_data_mytable,data_module, sheet_name, data_wert_new, data_wert, wert_pos_col):
    if data_wert_new:
        print(f"type data_wert_new: {type(data_wert_new)}")
        logger.debug(f"type data_wert_new: {type(data_wert_new)}")
        workbook_write = openpyxl.load_workbook(file_data_mytable)
        worksheet_write = workbook_write[sheet_name]
        if data_wert:
            print(f"type data_wert: {type(data_wert)}")
            logger.debug(f"type data_wert: {type(data_wert)}")
            if type(data_wert_new) == str and type(data_wert) == str:
                if simp_str(data_wert_new) != simp_str(data_wert):
                    worksheet_write.cell(data_module.position[0] + 1, wert_pos_col + 1,
                                         data_wert_new[0:(len(data_wert_new) - 1)])
                    worksheet_write.cell(data_module.position[0] + 1, wert_pos_col + 1).fill = PatternFill(
                        patternType='solid', fgColor='FFFF00')
                    workbook_write.save(filename=file_data_mytable)
                    print(f"保存Excel")
                    logger.info(f"保存Excel")
            if type(data_wert_new) == str and type(data_wert) != str:
                try:
                    if float(data_wert_new) != data_wert:
                        worksheet_write.cell(data_module.position[0] + 1, wert_pos_col + 1,
                                             data_wert_new[0:(len(data_wert_new) - 1)])
                        worksheet_write.cell(data_module.position[0] + 1, wert_pos_col + 1).fill = PatternFill(
                            patternType='solid', fgColor='FFFF00')
                        """
                                白色：FFFFFF，黑色：000000，红色：FF0000，黄色：FFFF00
                                绿色：00FF00，蓝色：0000FF，橙色：FF9900，灰色：C0C0C0
                                常见颜色代码表：https://www.osgeo.cn/openpyxl/styles.html#indexed-colours
                            """
                        workbook_write.save(filename=file_data_mytable)
                        print(f"保存Excel")
                        logger.info(f"保存Excel")
                except ValueError:
                    worksheet_write.cell(data_module.position[0] + 1, wert_pos_col + 1,
                                         data_wert_new[0:(len(data_wert_new) - 1)])
                    worksheet_write.cell(data_module.position[0] + 1, wert_pos_col + 1).fill = PatternFill(
                        patternType='solid', fgColor='FFFF00')
                    workbook_write.save(filename=file_data_mytable)
                    print(f"保存Excel")
                    logger.info(f"保存Excel")


        else:
            worksheet_write.cell(data_module.position[0] + 1, wert_pos_col + 1,
                                 data_wert_new[0:(len(data_wert_new) - 1)])
            worksheet_write.cell(data_module.position[0] + 1, wert_pos_col + 1).fill = PatternFill(
                patternType='solid', fgColor='FFFF00')
            workbook_write.save(filename=file_data_mytable)
            print(f"保存Excel")
            logger.info(f"保存Excel")


def main_data_query(file_a2l,file_geskon,file_dcm,file_data_mytable,file_mech_table):
    print(f"Debug:{file_data_mytable}")
    logger.info(f"Data表为:{file_data_mytable}")
    workbook = xlrd.open_workbook(file_data_mytable)
    list_sheet_names = workbook.sheet_names()
    print(f"list_sheet_names = {list_sheet_names}")
    logger.debug(f"list_sheet_names = {list_sheet_names}")
    list_data_object =[]
    for sheet_name in list_sheet_names:
        print(f"sheet_name = {sheet_name}")
        logger.debug(f"sheet_name = {sheet_name}")
        mytable = workbook.sheet_by_name(sheet_name)
        first_row = mytable.row_values(0, 0, None)
        print(f"first_row = {first_row}")
        logger.debug(f"first_row = {first_row}")
        check_title_result, data_module_pos_col, wert_pos_col, tuning_pra_pos_col = check_title(first_row)
        if check_title_result:
            data_module_col = mytable.col_values(data_module_pos_col, 0, None)
            i = 1
            for data_module_name in data_module_col:
                data_module = DataModule(sheet_name, [i, data_module_pos_col], data_module_col[i],file_a2l,file_geskon,file_dcm,file_data_mytable,file_mech_table)
                data_name = data_module.get_name(mytable)
                data_variable = data_module.get_variable(mytable, tuning_pra_pos_col)
                data_wert = data_module.get_wert(mytable, wert_pos_col)
                data_wert_new = data_module.get_wert_new(file_a2l,file_geskon,file_dcm,file_data_mytable,file_mech_table)
                i += 1
                print(f"i = {i}")
                print(
                    f"dataModule: {data_module.sheet},\n"
                    f"postion: {data_module.position},\n"
                    f"name: {data_module.name},\n"
                    f"wert: {data_module.wert},\n"
                    f"variable: {data_module.variable},\n"
                    f"new wert: {data_module.wert_new}\n---END---")
                logger.info(
                    f"功能名: {data_module.sheet},\n"
                    f"位置: {data_module.position},\n"
                    f"参数名: {data_module.name},\n"                    
                    f"变量名: {data_module.variable},\n"
                    f"原参数值: {data_module.wert},\n"
                    f"新参数值: {data_module.wert_new}")
                if data_module.wert_new != "":
                    list_data_object.append(data_module)
                print(f"list data module = /n{list_data_object}")

                # 更新写入wert_new
                update_wert(file_data_mytable,data_module, sheet_name, data_wert_new, data_wert, wert_pos_col)

                if i == mytable.nrows:
                    break

    from create_json import create_json_from_data
    create_json_from_data(list_data_object)




class DataModule:
    sheet = ""
    position = ""  # postion[0] is row, postion[1] is col
    name = ""
    wert = ""
    variable = ""
    wert_new = ""

    def __init__(self, sheet, position, name,file_a2l,file_geskon,file_dcm,file_data_mytable,file_mech_table):
        self.sheet = sheet
        self.position = position
        self.name = name

    def get_name(self, mytable):
        print(f"sheet = {self.sheet}")
        logger.debug(f"sheet = {self.sheet}")
        print(f"position = {self.position}")
        logger.debug(f"position = {self.position}")
        print(f"name = {self.name}")
        logger.debug(f"name = {self.name}")
        print(f"data_name = {mytable.cell_value(self.position[0], self.position[1])}")
        logger.debug(f"data_name = {mytable.cell_value(self.position[0], self.position[1])}")
        self.name = mytable.cell_value(self.position[0], self.position[1])
        return self.name

    def get_variable(self, mytable, tuning_pra_pos_col):
        # data_variable = mytable.cell_value((self.position[0]), self.position[1]+3)
        data_variable = mytable.cell_value((self.position[0]), tuning_pra_pos_col).replace("\u200b", "?")
        print(f"data_variable = {data_variable}")
        logger.debug(f"data_variable = {data_variable}")
        self.variable = data_variable
        print(f"self.variable = {self.variable}")
        logger.debug(f"self.variable = {self.variable}")
        return self.variable

    def get_wert(self, mytable, wert_pos_col):
        # data_wert = mytable.cell_value(self.position[0], self.position[1]+1)
        data_wert = mytable.cell_value(self.position[0], wert_pos_col)
        print(f"data_wert = {data_wert}")
        logger.debug(f"data_wert = {data_wert}")
        self.wert = data_wert
        return self.wert

    def get_wert_new(self,file_a2l,file_geskon,file_dcm,file_data_mytable,file_mech_table):
        # find_result, paragraph_find, = fun_find_text(self.variable, file_geskon, file_dcm)
        find_result, paragraph_find, = fun_find_para(self.name, self.variable,file_a2l,file_geskon,file_dcm,file_data_mytable,file_mech_table)
        if not find_result:
            self.wert_new = ""
        else:
            paragraph_find = format_paragraph_find(paragraph_find)
            self.wert_new = paragraph_find
        print(f"wert_new = {self.wert_new}")
        logger.info(f"获取新参数值:\n{self.wert_new}")
        return self.wert_new

def main(file_data_mytable):
    # file_a2l = get_file_path.file_a2l
    # file_dcm = get_file_path.file_dcm
    # file_geskon = get_file_path.file_geskon
    # file_mech_table = get_file_path.file_mech_table
    # file_data_mytable = get_file_path.file_data_mytable

    # file_a2l, file_geskon, file_dcm, file_data_mytable, file_mech_table = get_file_path.get_file_path()
    # print(file_a2l)
    # print(file_geskon)
    # print(file_dcm)
    # print(file_data_mytable)
    # print(file_mech_table)
    # file_a2l = "D:\MyPython\myProjects\DF0D01C0100_RG3_X_SCU3_B3_VAR_01.a2l"
    # file_geskon = "D:\MyPython\myProjects\XP1100D0100_RG3_X_SCU3_B3_VAR_01_geskon.kon"
    # file_dcm = "D:\MyPython\myProjects\DCM_ALL_XIAO PENG_F30_First Tuning(Normal mode_01+Comfortable_06+Sport_11)_Based on XP1100B0100_20220705.DCM"
    # file_data_mytable = "D:\MyPython\myProjects\DataQuery\data.xlsx"
    # file_mech_table = "D:\MyPython\myProjects\MechanicalDataSheet -机械参数表 Voyah_H97C_20220902.xlsm"
    main_data_query(file_data_mytable)

if __name__ == '__main__':
    t_start = time.perf_counter()
    main(file_data_mytable)
    t_end = time.perf_counter()
    t_cost = t_end - t_start

    print(f'运行耗时:{t_cost:.8f}s')
    logger.info(f'运行耗时:{t_cost:.8f}s')
    logger.info(f'进程结束')

    os.system("Pause")
